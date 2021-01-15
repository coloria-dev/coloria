"""
Helper tool for converting XLSX data to YAML.
"""
import numpy as np
import openpyxl
import yaml

wb = openpyxl.load_workbook("xiao.xlsx")

index = 3
filename = "unique_blue.yaml"

# first worksheet
ws = wb.worksheets[index]
session1_x = [str(ws[f"C{k}"].value) for k in range(4, 1669)]
session1_y = [str(ws[f"D{k}"].value) for k in range(4, 1669)]
session1_z = [str(ws[f"E{k}"].value) for k in range(4, 1669)]

session2_x = [str(ws[f"G{k}"].value) for k in range(4, 1669)]
session2_y = [str(ws[f"H{k}"].value) for k in range(4, 1669)]
session2_z = [str(ws[f"I{k}"].value) for k in range(4, 1669)]

session3_x = [str(ws[f"K{k}"].value) for k in range(4, 1669)]
session3_y = [str(ws[f"L{k}"].value) for k in range(4, 1669)]
session3_z = [str(ws[f"M{k}"].value) for k in range(4, 1669)]

data = np.array(
    [
        [session1_x, session2_x, session3_x],
        [session1_y, session2_y, session3_y],
        [session1_z, session2_z, session3_z],
    ]
).T

data = data.reshape(-1, 9, 3, 3)
data = np.moveaxis(data, 1, 2)

with open(filename, "w") as outfile:
    yaml.dump(data.tolist(), outfile)
